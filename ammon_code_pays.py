from pathlib import Path

import openpyxl


class PaysCode:

    def __init__(self, template_path=None):
        self.template_path = template_path
        self.pays_codes = self._load_pays_codes()

    def _load_pays_codes(self):
        """Charge les codes pays depuis l'onglet Pays du template"""
        pays_map = {'FRANCE': 'FRA', 'France': 'FRA'}

        if self.template_path and Path(self.template_path).exists():
            try:
                wb = openpyxl.load_workbook(self.template_path)
                if 'Pays' in wb.sheetnames:
                    ws_pays = wb['Pays']
                    for row in ws_pays.iter_rows(min_row=2, values_only=True):
                        if row[0] and row[1]:  # Code et Libellé
                            code = str(row[0]).strip()
                            libelle = str(row[1]).strip()
                            pays_map[libelle] = code
                            pays_map[libelle.upper()] = code
            except:
                pass

        return pays_map

    def get_pays_code(self, pays_libelle):
        """Retourne le code pays (ex: FRANCE → FRA)"""
        if not pays_libelle:
            return 'FRA'

        pays_libelle = pays_libelle.strip()

        if pays_libelle in self.pays_codes:
            return self.pays_codes[pays_libelle]

        for key, value in self.pays_codes.items():
            if key.upper() == pays_libelle.upper():
                return value

        return 'FRA'  # Par défaut
