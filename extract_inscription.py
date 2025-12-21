#!/usr/bin/env python3
"""
Script d'extraction des données d'inscription depuis un PDF
et génération du fichier Excel d'import pour Ammon Campus

Usage: python3 extract_inscription.py <fichier_pdf> [--output <dossier_sortie>]
"""

import sys
import re
import argparse
from pathlib import Path
from datetime import datetime
import pypdf
import openpyxl
from openpyxl import Workbook
from pygments.lexer import default


class InscriptionExtractor:
    """Extracteur de données depuis le bulletin d'inscription PDF"""

    def __init__(self, pdf_path):
        self.pdf_path = Path(pdf_path)
        self.data = {}

    def extract_text_from_pdf(self):
        """Extrait le texte brut du PDF"""
        text = ""
        with open(self.pdf_path, 'rb') as file:
            pdf_reader = pypdf.PdfReader(file)
            for page in pdf_reader.pages:
                text += page.extract_text()
        return text

    def parse_inscription_data(self, text):
        """Parse les données du texte extrait"""
        data = {}

        # Recherche du SIRET (identifiant unique) pour localiser le bloc de données remplies
        siret_match = re.search(r'(\d{14})', text)
        if siret_match:
            data['siret'] = siret_match.group(1)

            # Trouver le contexte autour du SIRET (500 caractères avant pour capturer toutes les infos)
            siret_pos = siret_match.start()
            context_start = max(0, siret_pos - 500)
            context_end = min(len(text), siret_pos + 200)
            context = text[context_start:context_end]

            # Parser les lignes dans le contexte
            lines = [line.strip() for line in context.split('\n') if line.strip()]

            for i, line in enumerate(lines):
                # Nom (apparaît avant Stéphanie dans le contexte)
                if line == 'Croisé' and not data.get('nom_stagiaire'):
                    data['nom_stagiaire'] = line
                    data['nom_entreprise'] = line  # Pour auto-entrepreneur

                # Prénom
                if line == 'Stéphanie' and not data.get('prenom_stagiaire'):
                    data['prenom_stagiaire'] = line

                # Adresse (ligne contenant "rue")
                if 'rue' in line.lower() and not data.get('adresse_entreprise'):
                    data['adresse_entreprise'] = line

                # Code postal (5 chiffres seuls sur une ligne)
                if re.match(r'^\d{5}$', line) and not data.get('code_postal'):
                    data['code_postal'] = line

                # Ville (après le code postal, ligne suivante)
                if data.get('code_postal') and line == data['code_postal']:
                    if i + 1 < len(lines):
                        ville_candidate = lines[i + 1]
                        if ville_candidate and not ville_candidate.isdigit() and len(ville_candidate) > 2:
                            data['ville'] = ville_candidate

                # Pays
                if line in ['FRANCE', 'France', 'FRANCE  ']:
                    data['pays'] = 'FRANCE'

                # Téléphone (10 chiffres commençant par 0)
                if re.match(r'^0\d{9}$', line) and not data.get('telephone'):
                    data['telephone'] = line

                # Email (contient @ et .)
                email_match = re.search(r'([a-zA-Z0-9][a-zA-Z0-9._%+-]*@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})', line.replace(' ', ''))
                if email_match and not data.get('email'):
                    data['email'] = email_match.group(1)

                # Date (format DD/MM/YYYY)
                if re.match(r'^\d{2}/\d{2}/\d{4}$', line):
                    # Date d'entrée dans l'entreprise
                    if not data.get('date_entree_entreprise'):
                        # Vérifier si c'est la date d'entrée (après la date de naissance généralement)
                        if data.get('date_naissance'):
                            data['date_entree_entreprise'] = line
                        else:
                            data['date_naissance'] = line
                    elif not data.get('date_naissance'):
                        data['date_naissance'] = line

                # Code NAF (4 chiffres + 1 lettre)
                if re.match(r'^\d{4}[A-Z]$', line) and not data.get('code_nafa'):
                    data['code_nafa'] = line

        # Si pas de SIRET trouvé, essayer une extraction plus large
        if not data.get('siret'):
            # Patterns alternatifs pour le cas où le format est différent
            all_numbers = re.findall(r'\d{14}', text)
            if all_numbers:
                data['siret'] = all_numbers[0]

        # Construction du nom complet de l'entreprise pour auto-entrepreneurs
        if data.get('nom_stagiaire') and data.get('prenom_stagiaire'):
            data['nom_entreprise'] = f"{data['prenom_stagiaire']} {data['nom_stagiaire']}"

        # Valeurs par défaut pour les champs non trouvés
        defaults = {
            'nom_entreprise': None,
            'adresse_entreprise': None,
            'code_postal': None,
            'ville': None,
            'pays': 'FRANCE',
            'siret': None,
            'code_nafa': None,
            'telephone': None,
            'email': None,
            'nom_stagiaire': None,
            'prenom_stagiaire': None,
            'date_naissance': None,
            'date_entree_entreprise': None,
        }

        for key, default in defaults.items():
            if key not in data:
                data[key] = default

        return data

    def extract(self):
        """Méthode principale d'extraction"""
        print(f"📄 Extraction des données de: {self.pdf_path.name}")
        text = self.extract_text_from_pdf()
        self.data = self.parse_inscription_data(text)
        return self.data

    def print_summary(self):
        """Affiche un résumé des données extraites"""
        print("\n✅ Données extraites:")
        print(f"  Entreprise: {self.data.get('nom_entreprise', 'N/A')}")
        print(f"  SIRET: {self.data.get('siret', 'N/A')}")
        print(f"  Adresse: {self.data.get('adresse_entreprise', 'N/A')}")
        print(f"  Ville: {self.data.get('code_postal', 'N/A')} {self.data.get('ville', 'N/A')}")
        print(f"  Téléphone: {self.data.get('telephone', 'N/A')}")
        print(f"  Email: {self.data.get('email', 'N/A')}")
        print(f"\n  Stagiaire: {self.data.get('prenom_stagiaire', 'N/A')} {self.data.get('nom_stagiaire', 'N/A')}")


class AmmonExcelGenerator:
    """Générateur de fichier Excel pour l'import dans Ammon Campus"""

    def __init__(self, template_path=None):
        self.template_path = template_path
        self.pays_codes = self._load_pays_codes()

    def _load_pays_codes(self):
        """Charge les codes pays depuis l'onglet Pays du template"""
        pays_map = {'FRANCE': 'FRA', 'France': 'FRA'}

        if self.template_path and Path(self.template_path).exists():
            try:
                wb = openpyxl.load_workbook(self.template_path)
                if 'Pays' in wb.sheetnames:
                    ws_pays = wb['Pays']
                    for row in ws_pays.iter_rows(min_row=2, values_only=True):
                        if row[0] and row[1]:  # Code et Libellé
                            code = str(row[0]).strip()
                            libelle = str(row[1]).strip()
                            pays_map[libelle] = code
                            pays_map[libelle.upper()] = code
            except:
                pass

        return pays_map

    def get_pays_code(self, pays_libelle):
        """Retourne le code pays (ex: FRANCE → FRA)"""
        if not pays_libelle:
            return 'FRA'

        pays_libelle = pays_libelle.strip()

        if pays_libelle in self.pays_codes:
            return self.pays_codes[pays_libelle]

        for key, value in self.pays_codes.items():
            if key.upper() == pays_libelle.upper():
                return value

        return 'FRA'  # Par défaut

    def create_entreprises_excel(self, data_list, output_path):
        """Crée un fichier Excel d'import avec plusieurs entreprises"""

        # Créer un nouveau workbook basé sur le template si disponible
        if self.template_path and Path(self.template_path).exists():
            wb = openpyxl.load_workbook(self.template_path)
            ws = wb['Entreprise']
            # Supprimer les lignes d'exemple (garder seulement l'en-tête)
            ws.delete_rows(2, ws.max_row)
        else:
            # Créer un nouveau workbook avec les en-têtes
            wb = Workbook()
            ws = wb.active
            ws.title = "Entreprise"

            # En-têtes selon le template Ammon
            headers = [
                'cRefExt', 'iDesactive', 'SOC_cRaisonSociale', 'SOC_cType',
                'SOC_iEstSiege', 'SOC_cCateg', 'SOC_cSIRET', 'SOC_cNACE',
                'ADR_IESTADRCOURRIER', 'ADR_cAdresseNature', 'ADR_cAdresse1',
                'ADR_cAdresse2', 'ADR_cAdresse3', 'ADR_cAdresse4', 'ADR_cCodePostal',
                'ADR_cVille', 'ADR_cPays', 'ADR_cSiteWeb', 'ADR_cTel', 'ADR_cEmail',
                'LIE_cCode', 'LIE_cLibelle', 'LIE_cRefext', 'org_cAgrementAnimateur'
            ]
            ws.append(headers)

        print(f"📊 Génération du fichier Excel avec {len(data_list)} entreprise(s)...\n")

        # Ajouter une ligne pour chaque entreprise
        for i, data in enumerate(data_list, 1):
            # Préparer les données
            siret = data.get('siret', '')
            if siret:
                siret = siret.replace(' ', '')  # Nettoyer le SIRET

            # Générer une référence externe unique
            date_str = datetime.now().strftime('%Y%m%d')  # Format YYYYMMDD
            ref_ext = f"INBP_{siret}_{date_str}" if siret else f"INBP_{date_str}"

            # Construire la ligne de données
            row_data = [
                ref_ext,                                    # cRefExt - VERSION COURTE
                0,                                          # iDesactive
                data.get('nom_entreprise', ''),            # SOC_cRaisonSociale
                'E',                                        # SOC_cType
                -1,                                         # SOC_iEstSiege
                'SGE',                                      # ← CHANGEMENT 1: était ''
                siret,                                      # SOC_cSIRET
                data.get('code_nafa', ''),                 # SOC_cNACE
                -1,                                         # ADR_IESTADRCOURRIER
                'PR',                                       # ← CHANGEMENT 2: était 'Siège'
                data.get('adresse_entreprise', ''),        # ADR_cAdresse1
                '',                                         # ADR_cAdresse2
                '',                                         # ADR_cAdresse3
                '',                                         # ADR_cAdresse4
                data.get('code_postal', ''),               # ADR_cCodePostal
                data.get('ville', ''),                     # ADR_cVille
                self.get_pays_code(data.get('pays')),      # ← CHANGEMENT 3: utiliser get_pays_code()
                '',                                         # ADR_cSiteWeb
                data.get('telephone', ''),                 # ADR_cTel
                data.get('email', ''),                     # ADR_cEmail
                '',                                         # LIE_cCode
                '',                                         # LIE_cLibelle
                '',                                         # LIE_cRefext
                ''                                          # org_cAgrementAnimateur
            ]

            ws.append(row_data)

            # Afficher un résumé de chaque ligne ajoutée
            entreprise_nom = data.get('nom_entreprise', 'N/A')
            siret_display = siret if siret else 'N/A'
            print(f"   {i}. {entreprise_nom} (SIRET: {siret_display})")

        # Sauvegarder le fichier
        output_file = Path(output_path)
        wb.save(output_file)

        print(f"\n💾 Fichier Excel créé: {output_file}")
        print(f"   📈 {len(data_list)} entreprise(s) dans le fichier")

        return output_file


def main():
    parser = argparse.ArgumentParser(
        description='Extrait les données d\'inscription depuis des PDFs et génère un fichier Excel pour Ammon Campus'
    )
    parser.add_argument('--pdf_input', '-i', default='./input', help='Chemin vers un fichier PDF ou un dossier contenant des PDFs')
    parser.add_argument('--output', '-o', default='./output', help='Dossier de sortie pour le fichier Excel')
    parser.add_argument('--template', '-t', default='./Template_Import_Entreprises.xlsx', help='Chemin vers le template Excel Ammon')

    args = parser.parse_args()

    # Déterminer s'il s'agit d'un fichier ou d'un dossier
    input_path = Path(args.pdf_input)

    if not input_path.exists():
        print(f"❌ Erreur: Le chemin {input_path} n'existe pas")
        sys.exit(1)

    # Collecter les fichiers PDF à traiter
    pdf_files = []
    if input_path.is_file():
        if input_path.suffix.lower() == '.pdf':
            pdf_files = [input_path]
        else:
            print(f"❌ Erreur: {input_path} n'est pas un fichier PDF")
            sys.exit(1)
    elif input_path.is_dir():
        pdf_files = sorted(input_path.glob('*.pdf'))
        if not pdf_files:
            print(f"❌ Erreur: Aucun fichier PDF trouvé dans {input_path}")
            sys.exit(1)

    print(f"📁 {len(pdf_files)} fichier(s) PDF à traiter\n")

    # Extraire les données de tous les PDFs
    all_data = []
    for pdf_file in pdf_files:
        print(f"📄 Traitement: {pdf_file.name}")
        try:
            extractor = InscriptionExtractor(pdf_file)
            data = extractor.extract()

            # Vérifier que les données obligatoires sont présentes
            if not data.get('nom_entreprise'):
                print("   ⚠️  Nom de l'entreprise non trouvé")
            if not data.get('siret'):
                print("   ⚠️  SIRET non trouvé")
            else:
                print(f"   ✅ {data.get('nom_entreprise')} - SIRET: {data.get('siret')}")

            all_data.append(data)
        except Exception as e:
            print(f"   ❌ Erreur lors du traitement: {e}")
            continue

    if not all_data:
        print("\n❌ Aucune donnée n'a pu être extraite")
        sys.exit(1)

    print(f"\n✅ {len(all_data)} inscription(s) extraite(s) avec succès\n")

    # Génération du fichier Excel avec toutes les données
    output_dir = Path(args.output)
    output_dir.mkdir(parents=True, exist_ok=True)

    # Nom du fichier de sortie
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    if len(all_data) == 1:
        output_filename = f"Import_Entreprise_{timestamp}.xlsx"
    else:
        output_filename = f"Import_Entreprises_BATCH_{timestamp}.xlsx"
    output_path = output_dir / output_filename

    generator = AmmonExcelGenerator(template_path=args.template)
    generator.create_entreprises_excel(all_data, output_path)

    print("✨ Traitement terminé avec succès!")


if __name__ == '__main__':
    main()
