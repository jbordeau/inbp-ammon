#!/usr/bin/env python3
"""
Script d'extraction des données d'inscription depuis un PDF
et génération du fichier Excel d'import pour Ammon Campus

Usage: python3 extract_inscription.py <fichier_pdf> [--output <dossier_sortie>]
"""

import sys
import argparse
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl import Workbook
import base64
import json
from mistralai import Mistral

api_key = "imHoWyGyaX0hHFhyHYvA5rc1IuRsE9Ob"
client = Mistral(api_key=api_key)


class InscriptionExtractor:
    """Extracteur de données depuis le bulletin d'inscription PDF"""

    def __init__(self, pdf_path):
        self.pdf_path = Path(pdf_path)
        self.data = {}

    def encode_file(self):
        with open(self.pdf_path, 'rb') as file:
            return base64.b64encode(file.read()).decode('utf-8')

    def call_mistral(self):
        base64_file = self.encode_file()
        return client.ocr.process(
        model="mistral-ocr-latest",
        pages=[0],
        document={
            "type": "document_url",
            "document_url": f"data:application/pdf;base64,{base64_file}"
        },
        include_image_base64=False,
        extract_footer=False,
        extract_header=False,
        document_annotation_format={
            "type": "json_schema",
            "json_schema": {
                "name": "response_schema",
                "schema": {
                    "type": "object",
                    "title": "StructuredData",
                    "required": [
                        "Civilité",
                        "Nom du stagiaire",
                        "Prénom du stagiaire",
                        "Adresse du stagiaire",
                        "Code postal du stagiaire",
                        "Ville du stagiaire",
                        "Pays du stagiare",
                        "Portable du stagiaire",
                        "Email du stagiaire",
                        "Date de naissance",
                        "nom de l'entreprise",
                        "adresse de l'entreprise",
                        "Code postal",
                        "Ville",
                        "Pays",
                        "Date d'entrée dans l'entreprise",
                        "Tél",
                        "N° de SIRET",
                        "Code NAFA",
                        "Email"
                    ],
                    "properties": {
                        "Civilité": {
                            "type": "string"
                        },
                        "Nom du stagiare": {
                            "type": "string"
                        },
                        "Prénom du stagiaire": {
                            "type": "string"
                        },
                        "Afresse du stagiaire": {
                            "type": "string"
                        },
                        "Code postal du stagiaire": {
                            "type": "string"
                        },
                        "Ville du stagiaire": {
                            "type": "string"
                        },
                        "Pays du stagiaire": {
                            "type": "string"
                        },
                        "Portable du stagiaire": {
                            "type": "string"
                        },
                        "Email du stagiaire": {
                            "type": "string"
                        },
                        "Date de naissance": {
                            "type": "string"
                        },
                        "nom de l'entreprise": {
                            "type": "string"
                        },
                        "adresse de l'entreprise": {
                            "type": "string"
                        },
                        "Code postal": {
                            "type": "string"
                        },
                        "Ville": {
                            "type": "string",
                        },
                        "Pays": {
                            "type": "string",
                        },
                        "Date d'entrée dans l'entreprise": {
                            "type": "string",
                        },
                        "Tél": {
                            "type": "string",
                        },
                        "N° de SIRET": {
                            "type": "string",
                        },
                        "Code NAFA": {
                            "type": "string",
                        },
                        "Email": {
                            "type": "string",
                        },
                    }
                }
            }
        }
    )

    def extract(self):
        """Méthode principale d'extraction"""
        print(f"📄 Extraction des données de: {self.pdf_path.name}")
        ocr_response = self.call_mistral()
        print(ocr_response)
        self.data = json.loads(ocr_response.document_annotation)
        return self.data

class EntrepriseExcelGenerator:
    """Générateur de fichier Excel pour l'import dans Ammon Campus"""

    def __init__(self, template_path=None):
        self.template_path = template_path
        self.pays_codes = self._load_pays_codes()

    def _load_pays_codes(self):
        """Charge les codes pays depuis l'onglet Pays du template"""
        pays_map = {'FRANCE': 'FRA', 'France': 'FRA'}

        if self.template_path and Path(self.template_path).exists():
            try:
                wb = openpyxl.load_workbook(self.template_path)
                if 'Pays' in wb.sheetnames:
                    ws_pays = wb['Pays']
                    for row in ws_pays.iter_rows(min_row=2, values_only=True):
                        if row[0] and row[1]:  # Code et Libellé
                            code = str(row[0]).strip()
                            libelle = str(row[1]).strip()
                            pays_map[libelle] = code
                            pays_map[libelle.upper()] = code
            except:
                pass

        return pays_map

    def get_pays_code(self, pays_libelle):
        """Retourne le code pays (ex: FRANCE → FRA)"""
        if not pays_libelle:
            return 'FRA'

        pays_libelle = pays_libelle.strip()

        if pays_libelle in self.pays_codes:
            return self.pays_codes[pays_libelle]

        for key, value in self.pays_codes.items():
            if key.upper() == pays_libelle.upper():
                return value

        return 'FRA'  # Par défaut

    def create_entreprises_excel(self, data_list, output_path):
        """Crée un fichier Excel d'import avec plusieurs entreprises"""

        # Créer un nouveau workbook basé sur le template si disponible
        if self.template_path and Path(self.template_path).exists():
            wb = openpyxl.load_workbook(self.template_path)
            ws = wb['Entreprise']
            # Supprimer les lignes d'exemple (garder seulement l'en-tête)
            ws.delete_rows(2, ws.max_row)
        else:
            # Créer un nouveau workbook avec les en-têtes
            wb = Workbook()
            ws = wb.active
            ws.title = "Entreprise"

            # En-têtes selon le template Ammon
            headers = [
                'cRefExt', 'iDesactive', 'SOC_cRaisonSociale', 'SOC_cType',
                'SOC_iEstSiege', 'SOC_cCateg', 'SOC_cSIRET', 'SOC_cNACE',
                'ADR_IESTADRCOURRIER', 'ADR_cAdresseNature', 'ADR_cAdresse1',
                'ADR_cAdresse2', 'ADR_cAdresse3', 'ADR_cAdresse4', 'ADR_cCodePostal',
                'ADR_cVille', 'ADR_cPays', 'ADR_cSiteWeb', 'ADR_cTel', 'ADR_cEmail',
                'LIE_cCode', 'LIE_cLibelle', 'LIE_cRefext', 'org_cAgrementAnimateur'
            ]
            ws.append(headers)

        print(f"📊 Génération du fichier Excel avec {len(data_list)} entreprise(s)...\n")

        # Ajouter une ligne pour chaque entreprise
        for i, data in enumerate(data_list, 1):
            # Préparer les données
            siret = data.get('N° de SIRET', '')
            if siret:
                siret = siret.replace(' ', '')  # Nettoyer le SIRET

            # Générer une référence externe unique
            date_str = datetime.now().strftime('%Y%m%d')  # Format YYYYMMDD
            ref_ext = f"INBP_{siret}_{date_str}" if siret else f"INBP_{date_str}"

            # Construire la ligne de données
            row_data = [
                ref_ext,                                    # cRefExt - VERSION COURTE
                0,                                          # iDesactive
                data.get('nom de l\'entreprise', ''),            # SOC_cRaisonSociale
                'E',                                        # SOC_cType
                -1,                                         # SOC_iEstSiege
                'SGE',                                      # ← CHANGEMENT 1: était ''
                siret,                                      # SOC_cSIRET
                data.get('Code NAFA', ''),                 # SOC_cNACE
                -1,                                         # ADR_IESTADRCOURRIER
                'PR',                                       # ← CHANGEMENT 2: était 'Siège'
                data.get('adresse de l\'entreprise', ''),        # ADR_cAdresse1
                '',                                         # ADR_cAdresse2
                '',                                         # ADR_cAdresse3
                '',                                         # ADR_cAdresse4
                data.get('Code postal', ''),               # ADR_cCodePostal
                data.get('Ville', ''),                     # ADR_cVille
                self.get_pays_code(data.get('Pays')),      # ← CHANGEMENT 3: utiliser get_pays_code()
                '',                                         # ADR_cSiteWeb
                data.get('Tél', ''),                 # ADR_cTel
                data.get('Email', ''),                     # ADR_cEmail
                '',                                         # LIE_cCode
                '',                                         # LIE_cLibelle
                '',                                         # LIE_cRefext
                ''                                          # org_cAgrementAnimateur
            ]

            ws.append(row_data)

            # Afficher un résumé de chaque ligne ajoutée
            entreprise_nom = data.get('nom_entreprise', 'N/A')
            siret_display = siret if siret else 'N/A'
            print(f"   {i}. {entreprise_nom} (SIRET: {siret_display})")

        # Sauvegarder le fichier
        output_file = Path(output_path)
        wb.save(output_file)

        print(f"\n💾 Fichier Excel créé: {output_file}")
        print(f"   📈 {len(data_list)} entreprise(s) dans le fichier")

        return output_file


def main():
    parser = argparse.ArgumentParser(
        description='Extrait les données d\'inscription depuis des PDFs et génère un fichier Excel pour Ammon Campus'
    )
    parser.add_argument('--pdf_input', '-i', default='./input', help='Chemin vers un fichier PDF ou un dossier contenant des PDFs')
    parser.add_argument('--output', '-o', default='./output', help='Dossier de sortie pour le fichier Excel')
    parser.add_argument('--template', '-t', default='./Template_Import_Entreprises.xlsx', help='Chemin vers le template Excel Ammon')

    args = parser.parse_args()

    # Déterminer s'il s'agit d'un fichier ou d'un dossier
    input_path = Path(args.pdf_input)

    if not input_path.exists():
        print(f"❌ Erreur: Le chemin {input_path} n'existe pas")
        sys.exit(1)

    # Collecter les fichiers PDF à traiter
    pdf_files = []
    if input_path.is_file():
        if input_path.suffix.lower() == '.pdf':
            pdf_files = [input_path]
        else:
            print(f"❌ Erreur: {input_path} n'est pas un fichier PDF")
            sys.exit(1)
    elif input_path.is_dir():
        pdf_files = sorted(input_path.glob('*.pdf'))
        if not pdf_files:
            print(f"❌ Erreur: Aucun fichier PDF trouvé dans {input_path}")
            sys.exit(1)

    print(f"📁 {len(pdf_files)} fichier(s) PDF à traiter\n")

    # Extraire les données de tous les PDFs
    all_data = []
    for pdf_file in pdf_files:
        print(f"📄 Traitement: {pdf_file.name}")
        try:
            extractor = InscriptionExtractor(pdf_file)
            data = extractor.extract()

            # Vérifier que les données obligatoires sont présentes
            if not data.get('nom_entreprise'):
                print("   ⚠️  Nom de l'entreprise non trouvé")
            if not data.get('siret'):
                print("   ⚠️  SIRET non trouvé")
            else:
                print(f"   ✅ {data.get('nom_entreprise')} - SIRET: {data.get('siret')}")

            all_data.append(data)
        except Exception as e:
            print(f"   ❌ Erreur lors du traitement: {e}")
            continue

    if not all_data:
        print("\n❌ Aucune donnée n'a pu être extraite")
        sys.exit(1)

    print(f"\n✅ {len(all_data)} inscription(s) extraite(s) avec succès\n")

    # Génération du fichier Excel avec toutes les données
    output_dir = Path(args.output)
    output_dir.mkdir(parents=True, exist_ok=True)

    # Nom du fichier de sortie
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    if len(all_data) == 1:
        output_filename = f"Import_Entreprise_{timestamp}.xlsx"
    else:
        output_filename = f"Import_Entreprises_BATCH_{timestamp}.xlsx"
    output_path = output_dir / output_filename

    generator = EntrepriseExcelGenerator(template_path=args.template)
    generator.create_entreprises_excel(all_data, output_path)

    print("✨ Traitement terminé avec succès!")


if __name__ == '__main__':
    main()
